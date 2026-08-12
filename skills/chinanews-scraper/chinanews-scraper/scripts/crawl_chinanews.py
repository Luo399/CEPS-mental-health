#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
中国新闻网滚动新闻爬虫
支持单页和多页数据爬取
支持JSON和Excel两种输出格式
"""

import argparse
import json
import sys
import time
from typing import List, Dict, Optional

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError as e:
    print(f"错误: 缺少必要的依赖包: {e}", file=sys.stderr)
    print("请安装依赖: pip install requests beautifulsoup4 openpyxl", file=sys.stderr)
    sys.exit(1)


class ChinanewsScraper:
    """中国新闻网爬虫类"""
    
    BASE_URL = "https://www.chinanews.com"
    USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    
    def __init__(self, user_agent: Optional[str] = None):
        """
        初始化爬虫
        
        Args:
            user_agent: 自定义User-Agent，默认使用内置UA
        """
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': user_agent or self.USER_AGENT,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
        })
    
    def fetch_page(self, page_num: int) -> str:
        """
        获取指定页面的HTML内容
        
        Args:
            page_num: 页码
            
        Returns:
            HTML内容字符串
        """
        url = f"{self.BASE_URL}/scroll-news/news{page_num}.html"
        
        try:
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            response.encoding = 'utf-8'
            return response.text
        except requests.RequestException as e:
            raise Exception(f"获取第{page_num}页失败: {str(e)}")
    
    def parse_news_list(self, html: str, page_num: int) -> List[Dict]:
        """
        解析HTML页面，提取新闻列表
        
        Args:
            html: HTML内容
            page_num: 当前页码
            
        Returns:
            新闻列表，每条新闻为字典格式
        """
        soup = BeautifulSoup(html, 'html.parser')
        news_list = []
        
        # 查找新闻列表容器
        content_list = soup.find('div', class_='content_list')
        if not content_list:
            return news_list
        
        # 提取所有新闻项
        items = content_list.find_all('li')
        for item in items:
            # 跳过分隔线
            if item.get('class') and 'nocontent' in item['class']:
                continue
            
            try:
                # 提取栏目
                category_elem = item.find('div', class_='dd_lm')
                category = category_elem.get_text(strip=True) if category_elem else ""
                
                # 提取标题和链接
                title_elem = item.find('div', class_='dd_bt').find('a')
                title = title_elem.get_text(strip=True) if title_elem else ""
                link = title_elem.get('href', '') if title_elem else ""
                
                # 提取时间
                time_elem = item.find('div', class_='dd_time')
                news_time = time_elem.get_text(strip=True) if time_elem else ""
                
                # 拼接完整URL
                full_url = self.BASE_URL + link if link else ""
                
                news_list.append({
                    'category': category,
                    'title': title,
                    'link': full_url,
                    'time': news_time,
                    'page': page_num
                })
            except Exception as e:
                # 跳过解析失败的新闻项
                continue
        
        return news_list
    
    def scrape_page(self, page_num: int) -> List[Dict]:
        """
        爬取单页新闻
        
        Args:
            page_num: 页码
            
        Returns:
            新闻列表
        """
        html = self.fetch_page(page_num)
        news_list = self.parse_news_list(html, page_num)
        return news_list
    
    def scrape_multiple_pages(self, start_page: int, end_page: int) -> List[Dict]:
        """
        爬取多页新闻
        
        Args:
            start_page: 起始页码
            end_page: 结束页码
            
        Returns:
            所有页面的新闻列表
        """
        all_news = []
        
        for page_num in range(start_page, end_page + 1):
            try:
                news_list = self.scrape_page(page_num)
                all_news.extend(news_list)
                print(f"第{page_num}页: 成功爬取{len(news_list)}条新闻", file=sys.stderr)
                
                # 请求间隔，避免过于频繁
                if page_num < end_page:
                    time.sleep(1)
            except Exception as e:
                print(f"第{page_num}页爬取失败: {str(e)}", file=sys.stderr)
                continue
        
        return all_news
    
    def save_to_excel(self, news_list: List[Dict], filename: str):
        """
        保存新闻列表到Excel文件
        
        Args:
            news_list: 新闻列表
            filename: 输出文件名
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError as e:
            raise Exception(f"缺少必要的依赖包: {e}. 请安装: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "新闻列表"
        
        # 表头样式
        header_fill = PatternFill(start_color="C72A2A", end_color="C72A2A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        headers = ['栏目', '标题', '链接', '时间', '页码']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入数据
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for row_num, news in enumerate(news_list, 2):
            ws.cell(row=row_num, column=1, value=news.get('category', '')).alignment = data_alignment
            ws.cell(row=row_num, column=2, value=news.get('title', '')).alignment = data_alignment
            ws.cell(row=row_num, column=3, value=news.get('link', '')).alignment = data_alignment
            ws.cell(row=row_num, column=4, value=news.get('time', '')).alignment = data_alignment
            ws.cell(row=row_num, column=5, value=news.get('page', '')).alignment = Alignment(horizontal="center", vertical="center")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        
        # 保存文件
        wb.save(filename)
        print(f"数据已保存到: {filename}", file=sys.stderr)


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='中国新闻网滚动新闻爬虫')
    parser.add_argument('--page', type=int, help='爬取指定页码的新闻（单页模式）')
    parser.add_argument('--start-page', type=int, help='爬取起始页码（多页模式）')
    parser.add_argument('--end-page', type=int, help='爬取结束页码（多页模式）')
    parser.add_argument('--output', type=str, default='json', choices=['json', 'excel'], help='输出格式：json或excel（默认json）')
    parser.add_argument('--filename', type=str, help='输出文件名（仅Excel格式需要，默认chinanews_news.xlsx）')
    parser.add_argument('--user-agent', type=str, help='自定义User-Agent')
    
    args = parser.parse_args()
    
    # 参数校验
    if args.page and (args.start_page or args.end_page):
        print("错误: 单页模式(--page)和多页模式(--start-page/--end-page)不能同时使用", file=sys.stderr)
        sys.exit(1)
    
    if not args.page and not (args.start_page and args.end_page):
        print("错误: 必须指定单页(--page)或多页(--start-page --end-page)模式", file=sys.stderr)
        sys.exit(1)
    
    if args.start_page and args.end_page and args.start_page > args.end_page:
        print("错误: 起始页码不能大于结束页码", file=sys.stderr)
        sys.exit(1)
    
    # Excel格式需要文件名
    if args.output == 'excel':
        filename = args.filename or 'chinanews_news.xlsx'
    
    # 创建爬虫实例
    scraper = ChinanewsScraper(user_agent=args.user_agent)
    
    try:
        # 单页模式
        if args.page:
            news_list = scraper.scrape_page(args.page)
            print(f"单页爬取完成: 第{args.page}页，共{len(news_list)}条新闻", file=sys.stderr)
        
        # 多页模式
        else:
            news_list = scraper.scrape_multiple_pages(args.start_page, args.end_page)
            print(f"多页爬取完成: 第{args.start_page}-{args.end_page}页，共{len(news_list)}条新闻", file=sys.stderr)
        
        # 根据输出格式处理
        if args.output == 'excel':
            scraper.save_to_excel(news_list, filename)
        else:
            # 输出JSON数据到标准输出
            json.dump(news_list, sys.stdout, ensure_ascii=False, indent=2)
        
    except Exception as e:
        print(f"爬取失败: {str(e)}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
